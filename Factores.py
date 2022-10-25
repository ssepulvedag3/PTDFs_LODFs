import pandas as pd
import networkx as nx
import numpy as np
import itertools
import openpyxl as xl

try:
    #----------Lecturas de datos----------------
    feeder=pd.read_excel('Datos.xlsx', sheet_name='Líneas')
    Demanda=pd.read_excel('Datos.xlsx', sheet_name='Demanda')
    Phi=pd.read_excel('Datos.xlsx', sheet_name='Phi')
    Costos=pd.read_excel('Datos.xlsx', sheet_name='Costos')
    Generadores=pd.read_excel('Datos.xlsx', sheet_name='Generadores')
    Ejectuar=pd.read_excel('Datos.xlsx', sheet_name='Ejecución')

    #--------------------------------- Caso con distintas contingencias -------------------------

    HM=feeder[feeder['HM']!=0].index.tolist() #Lineas para el mantenimiento (añadir +1 ya que python cuenta desde cero)
    lineas_en_mantenimiento=sum(feeder['HM']!=0) #Número de líneas en mantenimiento
    bina=list(itertools.product(*[[1,0]]*lineas_en_mantenimiento)) 
    tab_bin = pd.DataFrame(bina,columns=list(np.where(feeder['HM']!=0)[0])) #Posibles combinaciones
    pt_dic={}
    ld_dic={}

    data_pt={'Topología':[],'Nodos':[],'Lineas':[],'PTDF':[]}
    tabla_pt=pd.DataFrame(data_pt)
    data_df={'Topología':[],'Linea_l':[],'Linea_m':[],'LODF':[]}
    tabla_df=pd.DataFrame(data_df)
    len_datapt=[0]
    len_datadf=[0]

    nodos=len(set.union(set(feeder['Nodo_i']),set(feeder['Nodo_j'])))
    lineas_totales=len(feeder)
    
    disponibilidad=[1]*len(tab_bin)
    for k in range(len(tab_bin)):
        li_dis=[]
        combi=list(tab_bin.iloc[k])
        for m in range(len(combi)):
            if combi[m]==0:
                li_dis+=[HM[m]]
            else:
                continue
        new_feeder=feeder.drop(li_dis)  
        l=len(new_feeder)
        n=len(set.union(set(new_feeder['Nodo_i']),set(new_feeder['Nodo_j'])))

        Bp=np.diag([1/(1j*feeder['X'][mq]) for mq in range(l)])#Matriz de susceptancias primivitva
        A=np.zeros((n,l))
        cn=0
        cl=0
        for ldw in range(l):
            From=feeder['Nodo_i'][ldw]
            To=feeder['Nodo_j'][ldw]
            for m in range(n):
                if From==m+1:
                    A[m,ldw]=1
                elif To==m+1:
                    A[m,ldw]=-1
                    
        if 0 in np.sum(abs(A),axis=1) or len(A)!=nodos:
            print('la combinación ', combi, 'deja un nodo aislado, por lo que no se toma en cuenta')
            disponibilidad[k]=0
            pt_dic[k]=np.zeros((len(A[0,:]),nodos))
            ld_dic[k]=np.zeros((len(A[0,:]),len(A[0,:])))
            len_datapt+=[len_datapt[-1]+len(new_feeder['Línea'])+2]
            len_datadf+=[len_datadf[-1]+len(new_feeder['Línea'])+2]
            continue
            
        #--Matriz Bbus--
        Bbus=A@Bp@A.T
        #--Formación matrix X--
        Bbus[:,0]=0
        Bbus[0,:]=0
        Bbus[0,0]=1
        
        X=np.linalg.inv(Bbus)
        X=abs(X)
        X[0,0]=0
        
        Al=A.T #Matriz de incidencia Rama-nodo
        Bd=abs(Bp)

        #-------------Cálculo factores de sensibilidad Rama-Nodo ---------------
        lodf=np.zeros((l,l))
        ptdf_rr=np.zeros((l,l))
        Inci=Al@np.identity(n)
        
        ptdf_rn=Bd@Al@X#ptdf rama nodo
        ptdf_rr=Bd@Al@X@Al.T#ptdf rama rama
 
        #--Cálculo lodfs--
        for s in range(l):
            for kl in range(l):
                if ptdf_rr[kl,kl]==1:
                    ptdf_rr[kl,kl]=0
                lodf[s,kl]=ptdf_rr[s,kl]*(1/(1-ptdf_rr[kl,kl]))
        
        np.fill_diagonal(lodf,0)    
        
        ptdf_nuevo=np.zeros((lineas_totales,nodos))

        for k5 in li_dis:
            ptdf_rn=np.insert(ptdf_rn, k5, np.zeros((1, nodos)), 0)
            lodf=np.insert(lodf, k5, np.zeros((1, len(new_feeder))), 0)
    
        for k6 in li_dis:
            lodf=np.insert(lodf, k6, np.zeros((1, lineas_totales)), 1)
            
            
        pt_dic[k]=ptdf_rn
        ld_dic[k]=lodf
        pt_nf,pt_nc=ptdf_rn.shape
        df_nf,df_nc=lodf.shape
        
        fil_pf=[]
        row_pf=list(feeder['Línea'])
        col_pf=list(range(1,nodos+1))*pt_nf
        for k1 in row_pf:
            fil_pf+=list(np.repeat(k1,pt_nc))
   
        fil_df=[]
        row_df=list(feeder['Línea'])
        col_df=list(feeder['Línea'])*df_nf
        for k2 in row_df:
            fil_df+=list(np.repeat(k2,df_nc))
           
        data_pt2={'Topología': [k]*len(fil_pf), 'Nodos':col_pf,'Lineas':fil_pf,'PTDF':list(ptdf_rn.flatten())}   
        data_df2={'Topología': [k]*len(fil_df),'Linea_l':col_df,'Linea_m':fil_df,'LODF':list(lodf.flatten())}
        tabla_pt2=pd.DataFrame(data_pt2)
        tabla_df2=pd.DataFrame(data_df2)
        tabla_pt=pd.concat([tabla_pt,tabla_pt2])
        tabla_df=pd.concat([tabla_df,tabla_df2])
        len_datapt+=[len_datapt[-1]+len(feeder['Línea'])+2]
        len_datadf+=[len_datadf[-1]+len(feeder['Línea'])+2]
        
        
            
        
    tab_bin.rename(columns = {'Name' : 'First Name', 'age' : 'Age'}, inplace = True)

    tab_bin.insert(0, "Topología", list(range(len(tab_bin))), True)
    tab_bin.insert(1,"Disponibilidad", disponibilidad, True)
    with pd.ExcelWriter('datos.xlsx') as writer:
        Ejectuar.to_excel(writer, sheet_name='Ejecución',index=False)
        Generadores.to_excel(writer, sheet_name='Generadores',index=False)
        Costos.to_excel(writer, sheet_name='Costos',index=False)
        Phi.to_excel(writer, sheet_name='Phi',index=False)
        Demanda.to_excel(writer, sheet_name='Demanda',index=False)
        feeder.to_excel(writer, sheet_name='Líneas',index=False)
        tabla_pt.to_excel(writer, sheet_name='PTDF',index=False)
        tabla_df.to_excel(writer, sheet_name='LODF',index=False)
        tab_bin.to_excel(writer,sheet_name='Topología',index=False)
        
    with pd.ExcelWriter('PTDFS_LODFS.xlsx') as wroter:
        feeder.to_excel(wroter, sheet_name='lines',index=False)
        tab_bin.to_excel(wroter,sheet_name='Topologías',index=False)
        for k in pt_dic:
            df_ptdic=pd.DataFrame(pt_dic[k])
            df_lddic=pd.DataFrame(ld_dic[k])        
            df_ptdic.to_excel(wroter,sheet_name='PTDFS',startrow=len_datapt[k])
            df_lddic.to_excel(wroter,sheet_name='LODFS',startrow=len_datadf[k])


    #Ejecución exitosa 
    file = 'Datos.xlsx'
    book = xl.load_workbook(filename = file)
    sheetname = book.sheetnames
    sh = book['Ejecución']
    sh['A2'] = 1            #Bandera de cálculo de factores exitosos 
    book.save(file)      

except:
    file = 'Datos.xlsx'
    book = xl.load_workbook(filename = file)
    sheetname = book.sheetnames

    sh = book['Ejecución']
    sh['A2'] = -1           #Bandera de cálculo de factores con error

    #Borrar columnas
    sh = book['LODF']
    sh.delete_cols(1, sh.max_column)

    sh = book['PTDF']
    sh.delete_cols(1, sh.max_column)

    sh = book['Topología']
    sh.delete_cols(1, sh.max_column)

    colnames_LODF = ['Topología', 'Linea_l', 'Linea_m', 'LODF']
    LODF = pd.DataFrame(columns = colnames_LODF)

    colnames_PTDF = ['Topología', 'Nodos', 'Lineas', 'PTDF']
    PTDF = pd.DataFrame(columns = colnames_PTDF)

    colnames_topo = ['Topología', 'Disponibilidad']
    topo = pd.DataFrame(columns = colnames_topo)

    #Escribir hojas
    writer = pd.ExcelWriter(file, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title,ws) for ws in book.worksheets)
    LODF.to_excel(writer, 'LODF', index = False)
    PTDF.to_excel(writer, 'PTDF', index = False)
    topo.to_excel(writer, 'Topología', index = False)
    writer.save()
    book.save(file)